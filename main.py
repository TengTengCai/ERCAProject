import argparse

from openpyxl.reader.excel import load_workbook
from paddlenlp import Taskflow


class ReviewAnalysis(object):
    def __init__(self, data_file):
        self.data_file_path = data_file
        schema = [{"评价维度": ["观点词", "情感倾向[正向,负向,未提及]"]}]
        self.senta1 = Taskflow("sentiment_analysis", model="uie-senta-base", schema=schema)
        self.senta2 = Taskflow("sentiment_analysis", model="skep_ernie_1.0_large_ch")


    def load_data(self):
        wb = load_workbook(self.data_file_path)
        ws = wb.active
        mr = ws.max_row

        data = []
        data_list = ws[f"E2:E{mr}"]
        for row in data_list:
            for cell in row:
                data.append(cell.value)
        return data

    def run(self):
        pass


if __name__ == '__main__':
    parser = argparse.ArgumentParser()
    parser.add_argument("--file_path", default="datafile/B00S93EQUK-US-Reviews-20241202.xlsx", type=str, help="The data path of sentiment analysis.")
    args = parser.parse_args()

    ra = ReviewAnalysis(args.file_path)
    ra.run()
